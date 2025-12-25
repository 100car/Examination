"""
Утиліти для нормалізації та валідації артикулів (Symbol / SKU) у заданому діапазоні.

Що робить скрипт для кожної комірки у вказаному діапазоні:
1) Прибирає ВСІ пробільні символи (пробіли, таби, NBSP тощо) по всьому рядку.
2) Замінює візуально схожі кириличні літери на латинські (А->A, В->B, ...).
3) Перевіряє, що отриманий артикул існує в sku_master.xlsx:
   лист "SKU", колонка з заголовком "SKU (ключ, унікальний)".
4) Якщо не знайдено — пробує виправити через мапінг:
   sku_master.xlsx / лист "Mapping": "Неправильный" -> "Артикул".
5) Якщо все ще не знайдено:
   - підсвічує комірку червоним
   - записує значення + координати в лист "New" у звіті

Залежності: openpyxl
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Set, Tuple

import re

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------
# Візуальне підсвічування
# ---------------------------
FILL_RED = PatternFill(
    start_color="FFFF0000",
    end_color="FFFF0000",
    fill_type="solid",
)

FILL_CLEAR = PatternFill()  # "порожня" заливка (скидання)


# ---------------------------
# Нормалізація символів
# ---------------------------

# Мапа кирилиця -> латиниця для візуально схожих символів (верхній+нижній регістр).
# ВАЖЛИВО: нижній регістр замінюємо теж на ВЕЛИКІ латинські (A, B, ...),
# щоб артикул був уніфікований.
_C2L = {
    # Upper (кирилиця)
    "А": "A", "В": "B", "Е": "E", "І": "I", "К": "K", "М": "M",
    "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
    # Lower (кирилиця)
    "а": "A", "в": "B", "е": "E", "і": "I", "к": "K", "м": "M",
    "н": "H", "о": "O", "р": "P", "с": "C", "т": "T", "х": "X",
    # Додаткові “популярні” символи з практики
    "Ё": "E", "ё": "E",
    "Й": "I", "й": "I",
    "З": "3", "з": "3",
    "У": "Y", "у": "Y",
}

# Регулярка для будь-яких пробільних символів + NBSP і “вузьких” пробілів
_WS_RE = re.compile(r"[\s\u00A0\u2007\u202F]+")

# Назви листів/колонок в sku_master
SKU_SHEET_NAME = "SKU"
SKU_HEADER_NAME = "SKU (ключ, унікальний)"
MAPPING_SHEET_NAME = "Mapping"
MAPPING_WRONG_HEADER = "Неправильный"
MAPPING_RIGHT_HEADER = "Артикул"


@dataclass
class InvalidCell:
    """Опис проблемної комірки (після нормалізації/мапінгу артикул не знайдено)."""
    sheet: str
    address: str
    value: str


def normalize_symbol(value: object) -> str:
    """
    Нормалізує артикул:
    - None -> ""
    - trim
    - прибирає ВСІ пробіли у середині рядка (не тільки по краях)
    - замінює візуально схожі кириличні символи на латинські

    Повертає нормалізований рядок.
    """
    if value is None:
        return ""

    s = str(value).strip()
    if not s:
        return ""

    # Прибираємо пробіли “всередині” артикулу: "AB 12" -> "AB12"
    s = _WS_RE.sub("", s)

    # Замінюємо символи
    s = "".join(_C2L.get(ch, ch) for ch in s)
    return s


def _find_header_col(ws: Worksheet, header_name: str) -> Optional[int]:
    """
    Шукає колонку за назвою заголовка (в першому рядку).

    Повертає:
        номер колонки (1..N) або None.
    """
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is not None and str(v).strip() == header_name:
            return col
    return None


def _load_sku_and_mapping(sku_master_path: str) -> Tuple[Set[str], Dict[str, str]]:
    """
    Завантажує з sku_master.xlsx:
    - множину валідних SKU з листа SKU (колонка за заголовком SKU_HEADER_NAME)
    - словник мапінгу wrong->right з листа Mapping (колонки за заголовками)

    Важливо:
    - Нормалізуємо всі значення через normalize_symbol()
    """
    wb = openpyxl.load_workbook(sku_master_path, data_only=True)

    # ---- SKU set ----
    if SKU_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"У файлі sku_master немає листа '{SKU_SHEET_NAME}'")

    ws_sku = wb[SKU_SHEET_NAME]
    sku_col = _find_header_col(ws_sku, SKU_HEADER_NAME)
    if sku_col is None:
        raise ValueError(
            f"На листі '{SKU_SHEET_NAME}' не знайдено заголовок колонки '{SKU_HEADER_NAME}' у 1-му рядку"
        )

    sku_set: Set[str] = set()
    for r in range(2, ws_sku.max_row + 1):
        v = normalize_symbol(ws_sku.cell(r, sku_col).value)
        if v:
            sku_set.add(v)

    # ---- Mapping dict ----
    mapping: Dict[str, str] = {}
    if MAPPING_SHEET_NAME in wb.sheetnames:
        ws_map = wb[MAPPING_SHEET_NAME]

        wrong_col = _find_header_col(ws_map, MAPPING_WRONG_HEADER)
        right_col = _find_header_col(ws_map, MAPPING_RIGHT_HEADER)

        if wrong_col is None or right_col is None:
            # Якщо лист є, але заголовки не ті — краще явно сказати
            raise ValueError(
                f"На листі '{MAPPING_SHEET_NAME}' очікуються заголовки "
                f"'{MAPPING_WRONG_HEADER}' та '{MAPPING_RIGHT_HEADER}' у 1-му рядку"
            )

        for r in range(2, ws_map.max_row + 1):
            wrong = normalize_symbol(ws_map.cell(r, wrong_col).value)
            right = normalize_symbol(ws_map.cell(r, right_col).value)
            if wrong and right:
                mapping[wrong] = right

    return sku_set, mapping


def _prepare_new_sheet(wb: openpyxl.Workbook, new_sheet_name: str) -> Worksheet:
    """
    Готує лист для фіксації невідомих/помилкових артикулів.

    Якщо лист вже існує:
      - ми “перезаписуємо” заголовок і починаємо писати з рядка 2.
      - (Старі дані можна стерти повністю, але це повільно. Простий варіант:
         просто перезаписати зверху вниз новим вмістом.)

    Повертає об’єкт Worksheet.
    """
    if new_sheet_name in wb.sheetnames:
        ws_new = wb[new_sheet_name]
        # Мінімальне "очищення": заново ставимо заголовки,
        # а далі будемо перезаписувати рядки починаючи з 2.
        # Якщо хочеш повністю чистити лист, скажи — зроблю окремою опцією.
    else:
        ws_new = wb.create_sheet(new_sheet_name)

    ws_new["A1"].value = "Value"
    ws_new["B1"].value = "Sheet"
    ws_new["C1"].value = "Cell"
    return ws_new


def validate_symbol_range(
    report_path: str,
    symbol_range: str,
    sku_master_path: str,
    sheet_name: Optional[str] = None,
    new_sheet_name: str = "New",
    save_as: Optional[str] = None,
) -> List[InvalidCell]:
    """
    Нормалізує та перевіряє артикул у заданому діапазоні звіту.

    Аргументи:
        report_path: шлях до Excel-звіту, який треба обробити.
        symbol_range: Excel-діапазон, напр. "A2:A200" або "B:B" або "A1".
        sku_master_path: шлях до sku_master.xlsx.
        sheet_name: назва листа у звіті; якщо None — береться активний лист.
        new_sheet_name: назва листа, куди записувати нові/помилкові артикули.
        save_as: якщо задано — зберігаємо результат в інший файл (не перезаписуємо report_path).

    Повертає:
        список проблемних комірок (InvalidCell).
    """
    sku_set, mapping = _load_sku_and_mapping(sku_master_path)

    wb = openpyxl.load_workbook(report_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    ws_new = _prepare_new_sheet(wb, new_sheet_name)

    invalid: List[InvalidCell] = []
    out_row = 2

    for row in ws[symbol_range]:
        for cell in row:
            original = cell.value

            # 1) Нормалізуємо і записуємо назад у звіт
            val = normalize_symbol(original)
            cell.value = val

            # 2) Скидаємо заливку, якщо вона була
            cell.fill = FILL_CLEAR

            # Порожні значення не перевіряємо
            if not val:
                continue

            # 3) Перевіряємо наявність у майстер-списку
            if val not in sku_set:
                # 4) Пробуємо виправити через мапінг
                mapped = mapping.get(val)
                if mapped and mapped in sku_set:
                    cell.value = mapped
                    val = mapped

            # 5) Якщо досі не знайшли — підсвічуємо і записуємо в "New"
            if val not in sku_set:
                cell.fill = FILL_RED

                inv = InvalidCell(
                    sheet=ws.title,
                    address=cell.coordinate,
                    value=val,
                )
                invalid.append(inv)

                ws_new.cell(out_row, 1).value = val
                ws_new.cell(out_row, 2).value = ws.title
                ws_new.cell(out_row, 3).value = cell.coordinate
                out_row += 1

    out_path = save_as if save_as else report_path
    wb.save(out_path)
    return invalid


if __name__ == "__main__":
    """
    Приклад запуску з командного рядка:

    python process_symbol.py report.xlsx "A2:A200" sku_master.xlsx "Report" out.xlsx

    Де:
    - report.xlsx      — звіт
    - "A2:A200"        — діапазон
    - sku_master.xlsx  — майстер-файл
    - "Report"         — (опціонально) лист у звіті
    - out.xlsx         — (опціонально) куди зберегти (щоб не перезаписувати оригінал)
    """
    import sys

    if len(sys.argv) < 4:
        print(
            "Usage: python process_symbol.py <report.xlsx> <range> <sku_master.xlsx> "
            "[sheet_name] [save_as.xlsx]"
        )
        raise SystemExit(2)

    report, rng, master = sys.argv[1:4]
    sheet = sys.argv[4] if len(sys.argv) >= 5 else None
    save_as = sys.argv[5] if len(sys.argv) >= 6 else None

    bad = validate_symbol_range(
        report_path=report,
        symbol_range=rng,
        sku_master_path=master,
        sheet_name=sheet,
        save_as=save_as,
    )

    print(f"Готово. Некоректних комірок: {len(bad)}")
    for x in bad[:20]:
        print(f"- {x.sheet}!{x.address}: {x.value}")
