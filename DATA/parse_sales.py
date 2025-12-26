"""
Скрипт розпізнавання та "очищення" Excel-звітів ПРОДАЖІВ по регіонах (UA/KZ/UZ).

ОСНОВНІ ВИМОГИ (за твоїми правками):
1) Рядки без Symbol (артикула) — ВИДАЛЯЄМО (не переносимо у вихідний файл).
2) Заголовок "ОС" у рядку заголовків означає Stock.
   Якщо "ОС" присутній — регіон UZ визначається однозначно.
3) Спочатку відображаємо (unhide) ВСІ приховані стовпчики, потім обробляємо.
   Приховані колонки НЕ ігноруємо (включно в місяці та суму).

АЛГОРИТМ В ЦІЛОМУ:
- Визначаємо header row скорингом (Symbol + кількість місяців + Total/Stock).
- Symbol — обов’язково.
- Місяці — обов’язково (мінімум 1).
- Total/Stock — опціонально:
  - Total може бути відсутній -> ми все одно будуємо TOTAL як суму місяців.
  - Stock може бути відсутній -> створюємо порожню колонку Stock.
- Якщо оригінальний Total існує — підсвічуємо mismatch червоним.

Залежності: openpyxl
"""

from __future__ import annotations

import logging
import os
import re
import shutil
from dataclasses import dataclass
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


# =============================================================================
# ПАПКИ ПРОЄКТУ
# =============================================================================
SALES_REPORTS_DIR = "SALES_REPORTS"
RECOGNIZED_SALES_REPORTS_DIR = "RECOGNIZED_SALES_REPORTS"
FAILED_SALES_REPORTS_DIR = "FAILED_SALES_REPORTS"


# =============================================================================
# ЛОГІНГ
# =============================================================================
LOG_FILE = "sales_parser.log"


# =============================================================================
# СКАН ЗВЕРХУ ФАЙЛУ
# =============================================================================
PREFERRED_SHEET_NAME = "Arkusz1"
HEADER_SCAN_LIMIT = 150          # скільки рядків скануємо зверху
MAX_COLS_SCAN = 25               # скільки колонок скануємо зліва (ти казав <= 20)


# =============================================================================
# ДОПУСК ДЛЯ FLOAT
# =============================================================================
TOL = 1e-6


# =============================================================================
# КОРОТКИЙ ВИВІД В КОНСОЛІ
# =============================================================================
FILENAME_COL_WIDTH = 56


# =============================================================================
# ПІДСВІТКА MISMATCH
# =============================================================================
FILL_RED = PatternFill("solid", fgColor="FFC7CE")


# =============================================================================
# СИНОНІМИ КЛЮЧОВИХ КОЛОНОК (для м'якого contains-пошуку)
# =============================================================================
SYMBOL_NAMES_RAW = [
    "symbol",
    "sku",
    "item",
    "item no",
    "item number",
    "артикул",
    "арт",
    "код товара",
    "код",
    "кат номер",
    "кат. номер",
    "каталоговый номер",
    "каталожный номер",
    "номер каталога",
    "каталог",
]

TOTAL_NAMES_RAW = [
    "total",
    "всього",
    "всего",
    "итого",
    "разом",
    "sum",
    "suma",
    "grand total",
    "підсумок",
    "подсумок",
]

STOCK_NAMES_RAW = [
    "stock",
    "залишок",
    "залишки",
    "остаток",
    "остатки",
    "на складі",
    "на складе",
    "склад",
    "warehouse",
    "ending balance",
    "balance",
    "saldo",
    # ✅ ВАЖЛИВО: у тебе "ОС" = Stock
    "ос",
]


# =============================================================================
# НОРМАЛІЗАЦІЯ ТЕКСТУ
# =============================================================================
# Прибираємо пунктуацію, залишаємо букви/цифри/пробіли.
_PUNCT_RE = re.compile(r"[^0-9a-zа-яіїєёґ\s]+", flags=re.IGNORECASE)


def setup_logger() -> logging.Logger:
    """
    Налаштовує логер для детального логу в файл.
    В консоль логи не сиплемо — ти хочеш один рядок на файл.
    """
    logger = logging.getLogger("sales_parser")
    logger.setLevel(logging.DEBUG)

    # В Colab/ноутбуці можуть запускати багато разів — не дублюємо handlers
    if logger.handlers:
        return logger

    fmt = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger


LOGGER = setup_logger()


def ensure_dirs() -> None:
    """Створює потрібні папки, якщо їх немає."""
    os.makedirs(SALES_REPORTS_DIR, exist_ok=True)
    os.makedirs(RECOGNIZED_SALES_REPORTS_DIR, exist_ok=True)
    os.makedirs(FAILED_SALES_REPORTS_DIR, exist_ok=True)


def pick_sheet_name(wb: openpyxl.Workbook) -> str:
    """
    Обирає лист для обробки:
    - якщо Arkusz1 існує — беремо його
    - інакше — перший лист
    """
    return PREFERRED_SHEET_NAME if PREFERRED_SHEET_NAME in wb.sheetnames else wb.sheetnames[0]


def unmerge_all(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Розмерджує всі merged-комірки на листі та копіює значення верхньої-лівої
    комірки на весь блок.

    Це важливо для стабільного пошуку заголовків.
    """
    for merged_range in list(ws.merged_cells.ranges):
        min_r, min_c, max_r, max_c = merged_range.bounds
        top_left_value = ws.cell(min_r, min_c).value

        ws.unmerge_cells(str(merged_range))

        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                ws.cell(r, c).value = top_left_value


def unhide_all_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_cols: int = 300) -> None:
    """
    ✅ Вимога №3: "Спочатку відобрази всі приховані стовбчики"

    Робимо:
    - проходимо колонки 1..min(ws.max_column, max_cols)
    - ставимо hidden=False
    """
    max_col = min(ws.max_column, max_cols)
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].hidden = False


def normalize_text(value: object) -> str:
    """
    Нормалізує текст для порівняння заголовків:
    - None -> ""
    - lower
    - NBSP -> пробіл
    - прибирає пунктуацію
    - стискає пробіли
    """
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().lower()
    text = _PUNCT_RE.sub(" ", text)
    text = " ".join(text.split())
    return text


def build_keys(raw: Sequence[str]) -> List[str]:
    """
    Готує ключі для contains-match:
    - нормалізує
    - прибирає дублікати
    - сортує за довжиною (довші ключі першими)
    """
    uniq = {normalize_text(x) for x in raw if normalize_text(x)}
    return sorted(uniq, key=len, reverse=True)


SYMBOL_KEYS = build_keys(SYMBOL_NAMES_RAW)
TOTAL_KEYS = build_keys(TOTAL_NAMES_RAW)
STOCK_KEYS = build_keys(STOCK_NAMES_RAW)


def contains_any(text: str, keys: Sequence[str]) -> bool:
    """True, якщо text містить хоча б один key (substring)."""
    return bool(text) and any(k in text for k in keys)


def build_header_map(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    max_cols_scan: int = MAX_COLS_SCAN,
) -> Dict[int, str]:
    """
    Формує мапу заголовків {col_idx: normalized_header} у заданому рядку.
    """
    header_map: Dict[int, str] = {}
    max_col = min(ws.max_column, max_cols_scan)
    for c in range(1, max_col + 1):
        txt = normalize_text(ws.cell(header_row, c).value)
        if txt:
            header_map[c] = txt
    return header_map


def find_first_contains(header_map: Dict[int, str], keys: Sequence[str]) -> Optional[int]:
    """Знаходить перший col_idx, де заголовок містить один з keys."""
    for col_idx, header in header_map.items():
        if contains_any(header, keys):
            return col_idx
    return None


def as_number(value: object) -> float:
    """Перетворює value на float, якщо це число. Інакше 0.0"""
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def safe_copy_to_failed(path: str) -> str:
    """Копіює файл у FAILED папку та повертає шлях призначення."""
    dst = os.path.join(FAILED_SALES_REPORTS_DIR, os.path.basename(path))
    shutil.copy2(path, dst)
    return dst


# =============================================================================
# ДЕТЕКЦІЯ МІСЯЦІВ У ЗАГОЛОВКАХ (RU/UA/EN + числові)
# =============================================================================
RU_MONTHS = {
    1: ["январ"],
    2: ["феврал", "февр"],
    3: ["март"],
    4: ["апрел", "апр"],
    5: ["май"],
    6: ["июн"],
    7: ["июл"],
    8: ["август", "авг"],
    9: ["сентябр", "сен", "сент"],
    10: ["октябр", "окт"],
    11: ["ноябр", "ноя"],
    12: ["декабр", "дек"],
}

UA_MONTHS = {
    1: ["січ", "сiч"],
    2: ["лют"],
    3: ["берез", "бер"],
    4: ["квіт", "квит"],
    5: ["трав"],
    6: ["черв"],
    7: ["лип"],
    8: ["серп"],
    9: ["верес", "вер"],
    10: ["жовт"],
    11: ["листоп", "лист"],
    12: ["груд"],
}

EN_MONTHS = {
    1: ["jan", "january"],
    2: ["feb", "february"],
    3: ["mar", "march"],
    4: ["apr", "april"],
    5: ["may"],
    6: ["jun", "june"],
    7: ["jul", "july"],
    8: ["aug", "august"],
    9: ["sep", "sept", "september"],
    10: ["oct", "october"],
    11: ["nov", "november"],
    12: ["dec", "december"],
}

# Виявлення місяця у числовому форматі: 01..12 або 2025-11 тощо
MONTH_NUM_RE = re.compile(r"(?<!\d)(0?[1-9]|1[0-2])(?!\d)")


def month_index_from_header(header: str) -> Optional[int]:
    """
    Повертає номер місяця 1..12, якщо заголовок схожий на місяць.
    Інакше None.
    """
    if not header:
        return None

    h = header

    # Словесні місяці
    for idx, stems in RU_MONTHS.items():
        if any(stem in h for stem in stems):
            return idx
    for idx, stems in UA_MONTHS.items():
        if any(stem in h for stem in stems):
            return idx
    for idx, stems in EN_MONTHS.items():
        if any(stem in h for stem in stems):
            return idx

    # Числові місяці
    m = MONTH_NUM_RE.search(h)
    if m:
        num = int(m.group(1))
        if 1 <= num <= 12:
            return num

    return None


def detect_region(filename: str, header_map: Dict[int, str]) -> str:
    """
    Визначає регіон (UA/KZ/UZ/UNK) для префікса вихідного файлу.

    Пріоритет:
    1) Якщо у заголовках є "ос" (ОС) => Stock => регіон UZ однозначно.
    2) Маркери в назві файлу (UA/KZ/UZ).
    3) Визначення по мові місяців у заголовках.
    """
    # ✅ Вимога №2: "ОС" => UZ
    if any(h.strip() == "ос" for h in header_map.values()):
        return "UZ"

    up = filename.upper()
    if "UA" in up or "ЮА" in up:
        return "UA"
    if "KZ" in up:
        return "KZ"
    if "UZ" in up:
        return "UZ"

    text = " ".join(header_map.values())

    if any(stem in text for stems in UA_MONTHS.values() for stem in stems):
        return "UA"
    if any(stem in text for stems in RU_MONTHS.values() for stem in stems):
        return "KZ"
    if any(stem in text for stems in EN_MONTHS.values() for stem in stems):
        return "UZ"

    return "UNK"


# =============================================================================
# ПОШУК HEADER ROW СКОРИНГОМ
# =============================================================================
def score_row_as_header(header_map: Dict[int, str]) -> float:
    """
    Оцінює, наскільки рядок схожий на шапку.

    Бали:
    - Symbol присутній -> +10
    - кожна місячна колонка -> +2
    - Total -> +3
    - Stock -> +2
    - якщо дуже мало заголовків -> штраф
    """
    if not header_map:
        return -1.0

    has_symbol = find_first_contains(header_map, SYMBOL_KEYS) is not None
    has_total = find_first_contains(header_map, TOTAL_KEYS) is not None
    has_stock = find_first_contains(header_map, STOCK_KEYS) is not None

    month_count = sum(1 for h in header_map.values() if month_index_from_header(h) is not None)

    score = 0.0
    score += 10.0 if has_symbol else 0.0
    score += 2.0 * float(month_count)
    score += 3.0 if has_total else 0.0
    score += 2.0 if has_stock else 0.0

    if len(header_map) < 3:
        score -= 5.0

    return score


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """
    Знаходить header row у перших HEADER_SCAN_LIMIT рядках за максимальним score.
    """
    max_row = min(ws.max_row, HEADER_SCAN_LIMIT)

    best_row = 0
    best_score = -1.0

    for r in range(1, max_row + 1):
        hm = build_header_map(ws, r)
        sc = score_row_as_header(hm)
        if sc > best_score:
            best_score = sc
            best_row = r

    # Поріг: Symbol (10) + хоча б 1 місяць (2) + ще щось -> вже схоже на шапку
    if best_score < 12.0:
        raise ValueError("Не вдалося надійно знайти рядок заголовків (низький score).")

    LOGGER.debug("Header detection: best_row=%s best_score=%s", best_row, best_score)
    return best_row


# =============================================================================
# ДЕТЕКЦІЯ СТРУКТУРИ SALES-ЗВІТУ
# =============================================================================
@dataclass(frozen=True)
class SalesDetection:
    sheet_name: str
    header_row: int
    col_symbol: int
    month_cols: List[int]
    col_total: Optional[int]  # може бути None
    col_stock: Optional[int]  # може бути None
    region: str


def detect_sales_structure(
    ws_struct: openpyxl.worksheet.worksheet.Worksheet,
    filename: str,
) -> SalesDetection:
    """
    Визначає:
    - header_row
    - col_symbol (обов’язково)
    - month_cols (обов’язково)
    - col_total / col_stock (можуть бути None)
    - region
    """
    header_row = find_header_row(ws_struct)
    header_map = build_header_map(ws_struct, header_row)

    # Symbol — обов'язково
    col_symbol = find_first_contains(header_map, SYMBOL_KEYS)
    if col_symbol is None:
        raise ValueError("Не знайдено колонку Symbol/Артикул (обов’язкова).")

    # Total/Stock — опціонально
    col_total = find_first_contains(header_map, TOTAL_KEYS)

    # Stock: тут додаткова логіка для "ОС"
    col_stock = find_first_contains(header_map, STOCK_KEYS)
    if col_stock is None:
        # Шукаємо строго "ос"
        for col_idx, hdr in header_map.items():
            if hdr.strip() == "ос":
                col_stock = col_idx
                break

    # ✅ Місяці: беремо ВСІ колонки, що схожі на місяць.
    # ✅ НЕ пропускаємо hidden (бо ми їх уже показали unhide).
    month_pairs: List[Tuple[int, int]] = []
    for col_idx, hdr in header_map.items():
        m_idx = month_index_from_header(hdr)
        if m_idx is not None:
            month_pairs.append((m_idx, col_idx))

    month_pairs.sort(key=lambda x: x[0])
    month_cols = [col for _, col in month_pairs]

    if not month_cols:
        raise ValueError("Не знайдено жодної колонки місяця (обов’язково для sales-звіту).")

    region = detect_region(filename, header_map)

    LOGGER.debug(
        "Detected structure: header_row=%s symbol=%s months=%s total=%s stock=%s region=%s",
        header_row, col_symbol, month_cols, col_total, col_stock, region
    )

    return SalesDetection(
        sheet_name=ws_struct.title,
        header_row=header_row,
        col_symbol=col_symbol,
        month_cols=month_cols,
        col_total=col_total,
        col_stock=col_stock,
        region=region,
    )


# =============================================================================
# ОБРОБКА ОДНОГО ФАЙЛУ
# =============================================================================
def build_output_filename(original_filename: str, region: str) -> str:
    """UA_<base>_sales_recognized.xlsx"""
    base = original_filename[:-5] if original_filename.lower().endswith(".xlsx") else original_filename
    return f"{region}_{base}_sales_recognized.xlsx"


def clean_one_sales_file(input_path: str, output_path_tmp: str) -> Tuple[int, int, str]:
    """
    Обробляє один файл і зберігає у тимчасове ім’я (потім перейменуємо по region).

    Повертає:
        checked_rows: скільки рядків перенесли у вихід
        mismatches:   mismatch-рядки (лише якщо був оригінальний Total)
        region:       UA/KZ/UZ/UNK
    """
    filename = os.path.basename(input_path)
    LOGGER.info("START file=%s", input_path)

    wb_struct = openpyxl.load_workbook(input_path, data_only=False)
    sheet_name = pick_sheet_name(wb_struct)
    ws_struct = wb_struct[sheet_name]

    # ✅ Вимога №3: спочатку розмерджити, потім показати всі колонки
    unmerge_all(ws_struct)
    unhide_all_columns(ws_struct)

    wb_val = openpyxl.load_workbook(input_path, data_only=True)
    ws_val = wb_val[sheet_name]

    det = detect_sales_structure(ws_struct, filename)
    header_map = build_header_map(ws_struct, det.header_row)

    # Готуємо заголовки вихідного файлу:
    # Symbol + місяці (в тій послідовності, що визначили) + Stock + TOTAL
    out_headers: List[str] = ["Symbol"]
    for col_idx in det.month_cols:
        out_headers.append(header_map.get(col_idx, f"month_{col_idx}"))
    out_headers.append("Stock")
    out_headers.append("TOTAL")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = det.sheet_name

    # Запис header-рядка
    for j, h in enumerate(out_headers, start=1):
        out_ws.cell(1, j, h)

    checked = 0
    mismatches = 0
    out_row = 2

    # Ідемо по рядках ПІСЛЯ header row
    for r in range(det.header_row + 1, ws_struct.max_row + 1):
        if ws_struct.row_dimensions[r].hidden:
            continue

        # ✅ Вимога №1: якщо Symbol порожній — рядок ВИДАЛЯЄМО
        symbol_val = ws_val.cell(r, det.col_symbol).value
        symbol_txt = "" if symbol_val is None else str(symbol_val).strip()
        if not symbol_txt:
            continue

        # Місячні значення + computed_total
        month_values: List[float] = []
        computed_total = 0.0
        for c in det.month_cols:
            v = as_number(ws_val.cell(r, c).value)
            month_values.append(v)
            computed_total += v

        # Stock (якщо є)
        stock_val = ws_val.cell(r, det.col_stock).value if det.col_stock is not None else None

        # Порівняння з оригінальним Total (якщо він існує)
        if det.col_total is not None:
            original_total = as_number(ws_val.cell(r, det.col_total).value)
            row_mismatch = abs(computed_total - original_total) > TOL
            if row_mismatch:
                mismatches += 1
                LOGGER.debug(
                    "Mismatch row=%s computed_total=%s original_total=%s diff=%s",
                    r, computed_total, original_total, computed_total - original_total
                )
        else:
            row_mismatch = False

        # Запис у вихід:
        col_out = 1
        out_ws.cell(out_row, col_out, symbol_txt)
        col_out += 1

        for v in month_values:
            out_ws.cell(out_row, col_out, v)
            col_out += 1

        out_ws.cell(out_row, col_out, stock_val)
        col_out += 1

        out_ws.cell(out_row, col_out, computed_total)

        # Якщо mismatch — підсвітити рядок
        if row_mismatch:
            for c in range(1, len(out_headers) + 1):
                out_ws.cell(out_row, c).fill = FILL_RED

        checked += 1
        out_row += 1

    out_wb.save(output_path_tmp)
    LOGGER.info(
        "DONE file=%s region=%s checked=%s mismatches=%s saved=%s",
        input_path, det.region, checked, mismatches, output_path_tmp
    )
    return checked, mismatches, det.region


# =============================================================================
# BATCH-ОБРОБКА ПАПКИ
# =============================================================================
def main() -> None:
    """
    Обробляє всі .xlsx у SALES_REPORTS_DIR.

    Консоль: 1 рядок на файл
    Деталі: sales_parser.log
    """
    ensure_dirs()
    LOGGER.info("======== NEW RUN ========")

    files = [
        f for f in os.listdir(SALES_REPORTS_DIR)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]

    for filename in files:
        src = os.path.join(SALES_REPORTS_DIR, filename)

        # Тимчасове ім’я (region визначимо всередині)
        tmp_name = build_output_filename(filename, "UNK")
        tmp_dst = os.path.join(RECOGNIZED_SALES_REPORTS_DIR, tmp_name)

        try:
            checked, mism, region = clean_one_sales_file(src, tmp_dst)

            # Перейменування на фінальне ім'я з region-префіксом
            final_name = build_output_filename(filename, region)
            final_dst = os.path.join(RECOGNIZED_SALES_REPORTS_DIR, final_name)

            if os.path.abspath(final_dst) != os.path.abspath(tmp_dst):
                if os.path.exists(final_dst):
                    os.remove(final_dst)
                os.replace(tmp_dst, final_dst)
            else:
                final_dst = tmp_dst

            print(
                f"OK   {filename:<{FILENAME_COL_WIDTH}} | "
                f"region={region:<3} | checked={checked:<4} | mismatches={mism:<4} | "
                f"saved={os.path.basename(final_dst)}"
            )

        except Exception as exc:
            fail_dst = safe_copy_to_failed(src)
            LOGGER.exception("FAIL file=%s error=%s", src, exc)
            print(f"FAIL {filename} | {exc} | copied to {fail_dst}")


if __name__ == "__main__":
    main()
